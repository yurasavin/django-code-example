import io
import os
from datetime import date

from babel.dates import format_date
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment


def render_table(cleaned_data, worker):
    """
    Заполняет экселевский шаблон данными из формы и возвращает файл
    """
    template_path = os.path.join(
        '.',
        'nmck',
        'excel_template',
        'nmck_template.xlsx'
    )
    wb = load_workbook(template_path)
    ws = wb.active
    border_all = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
    )
    border_lftbold_other_simple = Border(
        left=Side(border_style='thick', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
    )
    border_rghtbold_other_simple = Border(
        right=Side(border_style='thick', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
    )
    border_bold_left = Border(
        left=Side(border_style='thick', color='FF000000'),
        top=Side(border_style='thick', color='FF000000'),
        bottom=Side(border_style='thick', color='FF000000'),
    )
    border_bold_right = Border(
        right=Side(border_style='thick', color='FF000000'),
        top=Side(border_style='thick', color='FF000000'),
        bottom=Side(border_style='thick', color='FF000000'),
    )
    border_up_down = Border(
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
    )
    border_up_down_right = Border(
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
    )
    alignment_hleft_vcenter_wrap = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=False,
    )
    alignment_hcenter_vcenter_wrap = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
        shrink_to_fit=False,
    )
    # Подсчет общей стоимости
    all_summ = 0
    # Записываем значения каждой позиции в строки
    for i, position in enumerate(cleaned_data, 9):
        if not position:
            i -= 1
            break
        # Номер ячейки
        ws[f'A{i}'] = i - 8
        ws[f'A{i}'].border = border_lftbold_other_simple
        ws[f'A{i}'].alignment = alignment_hcenter_vcenter_wrap

        # Имя позиции
        ws[f'B{i}'] = position['name']
        ws[f'B{i}'].border = border_all
        ws[f'B{i}'].alignment = alignment_hleft_vcenter_wrap

        # Единица измерения
        ws[f'C{i}'] = position['type_of_size']
        ws[f'C{i}'].border = border_all
        ws[f'C{i}'].alignment = alignment_hcenter_vcenter_wrap

        # Количество единиц
        ws[f'D{i}'] = position['num']
        ws[f'D{i}'].border = border_all
        ws[f'D{i}'].alignment = alignment_hcenter_vcenter_wrap

        # Цена 1
        ws[f'E{i}'] = position['price1']
        ws[f'E{i}'].border = border_all
        ws[f'E{i}'].alignment = alignment_hcenter_vcenter_wrap

        # Цена2
        ws[f'F{i}'] = position['price2']
        ws[f'F{i}'].border = border_all
        ws[f'F{i}'].alignment = alignment_hcenter_vcenter_wrap

        # Цена3
        ws[f'G{i}'] = position['price3']
        ws[f'G{i}'].border = border_all
        ws[f'G{i}'].alignment = alignment_hcenter_vcenter_wrap

        # Средняя цена
        ws[f'H{i}'] = position['avrg']
        ws[f'H{i}'].border = border_all
        ws[f'H{i}'].alignment = alignment_hcenter_vcenter_wrap

        # Коэф вариации
        ws[f'I{i}'] = position['variation']
        ws[f'I{i}'].border = border_all
        ws[f'I{i}'].alignment = alignment_hcenter_vcenter_wrap

        # Среднее квадратичное отклонение
        ws[f'J{i}'] = position['standard_deviation']
        ws[f'J{i}'].border = border_all
        ws[f'J{i}'].alignment = alignment_hcenter_vcenter_wrap

        # Рассчет цены
        ws[f'K{i}'] = '('
        ws[f'K{i}'].border = border_up_down
        ws[f'K{i}'].alignment = alignment_hcenter_vcenter_wrap

        ws[f'L{i}'] = position['price1']
        ws[f'L{i}'].border = border_up_down
        ws[f'L{i}'].alignment = alignment_hcenter_vcenter_wrap

        ws[f'M{i}'] = '+'
        ws[f'M{i}'].border = border_up_down
        ws[f'M{i}'].alignment = alignment_hcenter_vcenter_wrap

        ws[f'N{i}'] = position['price2']
        ws[f'N{i}'].border = border_up_down
        ws[f'N{i}'].alignment = alignment_hcenter_vcenter_wrap

        ws[f'O{i}'] = '+'
        ws[f'O{i}'].border = border_up_down
        ws[f'O{i}'].alignment = alignment_hcenter_vcenter_wrap

        ws[f'P{i}'] = position['price3']
        ws[f'P{i}'].border = border_up_down
        ws[f'P{i}'].alignment = alignment_hcenter_vcenter_wrap

        ws[f'Q{i}'] = ')/3*'
        ws[f'Q{i}'].border = border_up_down
        ws[f'Q{i}'].alignment = alignment_hcenter_vcenter_wrap

        ws[f'R{i}'] = position['num']
        ws[f'R{i}'].border = border_up_down
        ws[f'R{i}'].alignment = alignment_hcenter_vcenter_wrap

        ws[f'S{i}'] = '='
        ws[f'S{i}'].border = border_up_down
        ws[f'S{i}'].alignment = alignment_hcenter_vcenter_wrap

        ws[f'T{i}'] = position['summ']
        ws[f'T{i}'].border = border_rghtbold_other_simple
        ws[f'T{i}'].alignment = alignment_hcenter_vcenter_wrap

        all_summ += position['summ']
    i += 1
    ws[f'A{i}'] = 'Итого'
    ws.merge_cells(f'A{i}:S{i}')
    for cell in ws[f'A{i}:S{i}'][0]:
        cell.border = border_bold_left
    ws[f'A{i}'].alignment = alignment_hleft_vcenter_wrap

    ws[f'T{i}'] = all_summ
    ws[f'T{i}'].border = border_bold_right
    ws[f'T{i}'].alignment = alignment_hcenter_vcenter_wrap

    i += 1
    ws[f'A{i}'] = '* - коэффициент вариации не превышает 33 %, совокупность цен является однородной.'
    ws.merge_cells(f'A{i}:T{i}')

    i += 2
    ws[f'A{i}'] = 'Член контрактной службы:'
    ws.merge_cells(f'A{i}:C{i}')

    i += 1
    ws[f'A{i}'] = 'Начальник отдела закупок'
    ws.merge_cells(f'A{i}:C{i}')

    i += 1
    ws[f'A{i}'] = '_________ Савин Ю.К.'
    ws.row_dimensions[i].height = 30
    ws.merge_cells(f'A{i}:C{i}')

    i += 1
    ws[f'A{i}'] = format_date(date.today(), format='long', locale='ru_RU')
    ws.merge_cells(f'A{i}:C{i}')

    i += 1
    ws[f'A{i}'] = f'Исп. {worker}'
    ws.merge_cells(f'A{i}:C{i}')

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
