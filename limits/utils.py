"""
Модуль содержит класс ExcelPlanParser, который обновляет лимиты из файла ПФХД
"""
from decimal import ROUND_HALF_UP, Decimal

from django.db import transaction
from openpyxl import load_workbook

from limits.models import IndustryCode, Limit, LimitArticle, LimitMoney, Source


class ExcelPlanParser():
    """
    Парсит ПФХД в формате XLSX, обновляет лимиты в базе данных
    """
    # Года по которым будут обновляться лимиты
    YEARS = [
        2020,
        2021,
        2022,
    ]

    # Колонки в Екселе которым соответствуют суммы по году
    YEAR_COLUMNS = {
        2020: 'BG',
        2021: 'BR',
        2022: 'CB',
    }

    # Значение ячейки содержащей сумму лимитов по источнику
    TOTAL_NAME = ('Закупка товаров, работ и услуг для '
                  'обеспечения государственных (муниципальных) нужд')

    # Статьи которые надо найти
    ARTICLES = (
        '221',
        '222',
        '223',
        '224',
        '225',
        '226',
        '227',
        '228',
        '310',
        '341',
        '343',
        '344',
        '345',
        '346',
        '349',
    )

    # Статьи отраслевой код которых равен '00000'
    ARTICLES_WITH_ZERO_IND_CODE = (
        '224',
        '227',
        '228',
        '341',
        '343',
        '344',
        '345',
    )

    ARTICLE_NAMES = {
        '221': 'Услуги связи',
        '222': 'Транспортные услуги',
        '223': 'Коммунальные услуги',
        '224': 'Арендная плата за пользование имуществом',
        '225': 'Работы, услуги по содержанию имущества',
        '226': 'Прочие работы, услуги',
        '227': 'Страхование',
        '228': 'Услуги работы для целей капитальных вложений',
        '310': 'Увеличение стоимости основных средств',
        '341': ('Увеличение стоимости лекарственных препаратов и'
                ' материалов, применяемых в медицинских целях'),
        '343': 'Увеличение стоимости горюче-смазочных материалов',
        '344': 'Увеличение стоимости строительных материалов',
        '345': 'Увеличение стоимости мягкого инвентаря',
        '346': 'Увеличение стоимости прочих оборотных запасов (материалов)',
        '349': ('Увеличение стоимости прочих материальных '
                'запасов однократного применения'),
    }

    SOURCE_NAMES = {
        '2': 'Внебюджет',
        '4': 'Бюджет',
        '5': 'Целевые',
        '7': 'ОМС',
    }

    @classmethod
    def get_info(cls):
        """Возвращает информацию о статических переменных класса"""
        years = 'Года для которых обновляются лимиты: {}'.format(
            ', '.join((str(x) for x in cls.YEARS))
        )
        columns = 'Колонки в Екселе из которых берутся суммы: {}'.format(
            ', '.join(cls.YEAR_COLUMNS.values())
        )
        articles = 'Статьи по которым происходит поиск: {}'.format(
            ', '.join(cls.ARTICLES)
        )
        out_lines = [years, columns, articles]
        return out_lines

    def __init__(self, file):
        """Открывает файл и нужный лист в нем"""
        wb = load_workbook(file)
        self.sheet = wb.get_sheet_by_name(wb.sheetnames[1])
        self.response = ['Информация о ходе обновления лимитов:']
        self.limits = [
            Limit.objects.get_or_create(year=year) for year in self.YEARS
        ]
        self.source_money = None

    @transaction.atomic()
    def main(self):
        """Метод, управляющий логикой парсинга Эксель файла"""
        self.clean_limits()
        for limit in self.limits:
            self.limit = limit[0]
            self.money_column = self.YEAR_COLUMNS[self.limit.year]
            self.index = 0
            while self.set_index():
                self.parse_data()
                self.check_data()
                self.save_data()
        return self.response

    def clean_limits(self):
        """
        Обнуляет суммы в лимитах по которым планируется загрузить новые данные
        """
        for limit, created in self.limits:
            if not created:
                limit.set_money_to_zero()

    def get_index_value(self, column):
        """
        Возвращает строковое значение из книги в текущем индексе
        """
        return self.sheet[f'{column}{self.index}'].value

    def set_index(self):
        """
        Проверяет что значение в следующей ячейке не пустое,
        КВР входит в значения KVR,
        в случае успеха возвращает True меняет значение self.index,
        после 100 подряд пустых ячеек возвращает False
        Вывывает метод self.check_limits для проверки цифр
        """
        KVR = ('243', '244')
        missed_cells = 0
        while missed_cells < 100:
            self.index += 1
            value = self.get_index_value(column='BA')
            if not value:
                missed_cells += 1
                continue
            if self.get_index_value(column='BB') == '0':
                continue
            if self.get_index_value(column='AY') not in KVR:
                self.check_limits()
                missed_cells = 0
                continue
            return True
        return False

    def check_limits(self):
        """
        Проверяет наименование ячейки, если это сумма по источнику,
        то записывает в source_money, если в переменной уже есть значение,
        то перед сменой переменной проверяет сумму по текущему источнику
        """
        value = self.get_index_value(column='B')
        if not value:
            return
        value = ' '.join(value.split())
        if value == self.TOTAL_NAME:
            money = Decimal(self.get_index_value(self.money_column)).quantize(
                Decimal("1.00"), ROUND_HALF_UP)
        else:
            return
        if self.source_money:
            source_money_in_db = self.source.get_money()
            delta = source_money_in_db - self.source_money
            msg = '{}. Сумма в ПФХД: {} != напарсенному: {}'.format(
                self.source,
                money,
                source_money_in_db
            )
            assert delta == 0, msg
            to_resp = f'{self.limit.year} {self.source} - сумма сошлась'
            self.response.append(to_resp)
        self.source_money = money

    def parse_data(self):
        """Находит источник, статью, отраслевой код, сумму и соханяет их"""

        # Определяем источник
        self.source_num = self.get_index_value(column='BB')

        # Определяем статью
        value = self.get_index_value(column='BA')
        self.article_num = ''.join(value.split('.')[3:-1])

        # Определяем отраслевой код
        self.ind_code_num = value.split('.')[-1]

        # Находим сумму по коду
        value = self.get_index_value(column=self.money_column)
        self.money_value = Decimal(value).quantize(
            Decimal("1.00"), ROUND_HALF_UP)

        # Находим код субсидии
        self.sybsidy_code = self.get_index_value('AZ')

    def check_data(self):
        """Проверяет данные полученные из строки книги на соответствие"""
        data_is_checked = True

        # Если статья не из перечня статей пропускаем
        if self.article_num not in self.ARTICLES:
            data_is_checked = False

        # Если отраслевой код нулевой и статья не в списке статей с нулевым ок
        if (self.ind_code_num == '0000' and self.article_num not in
                self.ARTICLES_WITH_ZERO_IND_CODE):
            data_is_checked = False

        # Если источник целевые и код субсидии нулевой пропускаем
        if self.source_num == '5' and self.sybsidy_code == '00000':
            data_is_checked = False

        self.data_is_checked = data_is_checked

    def save_data(self):
        """
        Создает модели из напарсенного и сохраняет в БД,
        если data_is_checked == True
        """
        if not self.data_is_checked:
            return

        # Источник
        self.source, created = Source.objects.get_or_create(
            num=int(self.source_num),
            name=self.SOURCE_NAMES[self.source_num],
            limit=self.limit,
        )

        # Статья
        self.article, created = LimitArticle.objects.get_or_create(
            name=self.ARTICLE_NAMES[self.article_num],
            num=int(self.article_num),
            source=self.source,
        )

        # Отраслевой код
        self.industry_code, created = IndustryCode.objects.get_or_create(
            name=self.get_index_value(column='C'),
            num=self.ind_code_num,
            limit_article=self.article,
        )

        # Деньги
        name = 'Год: {}; {}; Статья: {}; Отраслевой код: {}'.format(
            self.limit.year,
            self.source.name,
            self.article.num,
            self.industry_code.num,
        )
        if self.sybsidy_code != '00000':
            name += f'; Код субсидии: {self.sybsidy_code}'
        self.money, created = LimitMoney.objects.get_or_create(
            name=name,
            sybsidy_code=self.sybsidy_code,
            industry_code=self.industry_code,
        )
        self.money.money = self.money_value
        self.money.save()

        to_response = f'Строка - {self.index}. Значение - {self.money}'
        self.response.append(to_response)
